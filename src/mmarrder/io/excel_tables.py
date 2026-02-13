from __future__ import annotations

from dataclasses import dataclass
from typing import Any, Mapping

import pandas as pd
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.utils import get_column_letter


@dataclass(frozen=True)
class ExcelTableExportOptions:
    engine: str = "openpyxl"
    table_style: str = "TableStyleLight13"
    autofit_columns: bool = True
    freeze_panes: str | None = "A2"


def export_to_excel_with_tables(
    dict_dataframes: Mapping[str, Any],
    file_name: str,
    *,
    options: ExcelTableExportOptions | None = None,
) -> str:
    """
    Exporta múltiples DataFrames a un .xlsx, cada uno en su hoja,
    y convierte el rango escrito en una Tabla estructurada de Excel (ListObject) con estilo.

    dict_dataframes puede ser:
      - {"Nombre Hoja": DataFrame}
      - {"Nombre Hoja": {"df": DataFrame, "table_name": "MiTabla"}}

    Retorna la ruta del archivo generado.
    """
    opts = options or ExcelTableExportOptions()

    def _sanitize_table_name(name: str) -> str:
        """
        Excel: nombre debe iniciar con letra/_, sin espacios, sin caracteres raros,
        sin "-", y único dentro del workbook.
        """
        import re

        name = (name or "").strip()
        name = re.sub(r"\s+", "_", name)
        name = re.sub(r"[^A-Za-z0-9_]", "_", name)
        if not name:
            name = "Tabla"
        if name[0].isdigit():
            name = f"T_{name}"
        return name[:255]

    def _autofit(ws, df: pd.DataFrame) -> None:
        for col_idx, col_name in enumerate(df.columns, start=1):
            if len(df):
                series = df[col_name].astype(str)
                max_len = max(len(str(col_name)), series.map(len).max())
            else:
                max_len = len(str(col_name))

            ws.column_dimensions[get_column_letter(col_idx)].width = min(
                max(10, int(max_len) + 2), 60
            )

    def _add_table(ws, df: pd.DataFrame, table_name: str) -> None:
        max_row = (len(df) + 1) if df is not None else 1
        max_col = len(df.columns) if df is not None else 1

        last_col_letter = get_column_letter(max_col)
        table_range = f"A1:{last_col_letter}{max_row}"

        table = Table(displayName=table_name, ref=table_range)
        style = TableStyleInfo(
            name=opts.table_style,
            showFirstColumn=False,
            showLastColumn=False,
            showRowStripes=True,
            showColumnStripes=False,
        )
        table.tableStyleInfo = style
        ws.add_table(table)

    used_table_names: set[str] = set()

    with pd.ExcelWriter(file_name, engine=opts.engine) as writer:
        for sheet_name, payload in dict_dataframes.items():
            if isinstance(payload, dict):
                df = payload.get("df")
                custom_table_name = payload.get("table_name")
            else:
                df = payload
                custom_table_name = None

            if df is None:
                raise ValueError(f"La hoja '{sheet_name}' no tiene DataFrame (df=None).")

            if not isinstance(df, pd.DataFrame):
                raise TypeError(
                    f"El valor de '{sheet_name}' debe ser un DataFrame o dict con 'df'. "
                    f"Recibido: {type(df)}"
                )

            # 1) escribir DF
            df.to_excel(writer, sheet_name=sheet_name, index=False)

            # 2) worksheet
            ws = writer.sheets[sheet_name]

            # 3) freeze panes
            if opts.freeze_panes:
                ws.freeze_panes = opts.freeze_panes

            # 4) nombre de tabla único
            base_name = custom_table_name or f"Tabla_{sheet_name}"
            base_name = _sanitize_table_name(base_name)

            table_name = base_name
            i = 1
            while table_name in used_table_names:
                i += 1
                suffix = f"_{i}"
                table_name = _sanitize_table_name(base_name[: (255 - len(suffix))] + suffix)

            used_table_names.add(table_name)

            # 5) tabla
            _add_table(ws, df, table_name)

            # 6) autofit
            if opts.autofit_columns:
                _autofit(ws, df)

    return file_name
